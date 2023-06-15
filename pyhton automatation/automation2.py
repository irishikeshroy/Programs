import openpyxl
import requests
from bs4 import BeautifulSoup
import re

# Define the Google search URL
GOOGLE_SEARCH_URL = 'https://www.google.com/search?q='

# Define the keyword to search for
KEYWORD = 'prepinsta'

# Define the maximum number of search result pages to analyze
MAX_PAGES = 10

# Open the input Excel file and get the first sheet
input_workbook = openpyxl.load_workbook('data.xlsx')
input_sheet = input_workbook.active

# Create a new output Excel file and worksheet
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active

# Write the headers to the output worksheet
output_sheet.append(['Statement', 'Page number', 'Link'])

# Loop through the rows in the input sheet, skipping the first row (header)
for row in input_sheet.iter_rows(min_row=2, values_only=True):
    # Get the search statement from the first column
    search_statement = row[0]
    
    # Create the Google search URL
    search_url = GOOGLE_SEARCH_URL + search_statement
    
    # Send a GET request to the search URL
    response = requests.get(search_url)
    
    # Parse the HTML content of the response with BeautifulSoup
    soup = BeautifulSoup(response.content, 'html.parser')
    
    # Find all the links in the search result pages
    links = soup.find_all('a')
    
    # Create a list to store the links and page numbers where the keyword is found
    found_links = []
    
    # Loop through the links and check if the keyword is found
    for i, link in enumerate(links):
        href = link.get('href')
        if href.startswith('/url?q='):
            url = re.search(r'/url\?q=(.+?)&', href).group(1)
            text = link.get_text()
            if KEYWORD.lower() in text.lower():
                found_links.append((url, i//10+1))
    
    # If at least one link with the keyword is found, store the search statement and links
    if found_links:
        for link, page_number in found_links:
            output_sheet.append([search_statement, page_number, link])
    # If no link with the keyword is found, store the search statement with a "not found" message
    else:
        output_sheet.append([search_statement, 'Not found', ''])
    
# Save the output workbook to a file
output_workbook.save('output_data.xlsx')
