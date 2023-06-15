import openpyxl
import requests
from bs4 import BeautifulSoup

# Load the Excel workbook
workbook = openpyxl.load_workbook('sheet1.xlsx')

# Select the first sheet
sheet1 = workbook['Sheet1']

# Select the second sheet
sheet2 = workbook['Sheet2']

# Set the limit for number of pages to be searched
limit = 30

# Loop through all rows in Sheet1
for i in range(1, sheet1.max_row + 1):
    statement = sheet1.cell(row=i, column=1).value
    print(f'Searching for: {statement}')

    # Perform a Google search
    response = requests.get(f'https://www.google.com/search?q={statement}')
    soup = BeautifulSoup(response.text, 'html.parser')
    text=soup.text
    

    # Loop through pages of Google search results
    for page in range(limit):
        links = soup.find_all("h3")
        found = False
        
        # Check if the keyword is present in the page
        for link in links:
            if "PrepInsta" in link.text:
                print(f'Found on page {page + 1}')
                sheet2.cell(row=i, column=1).value = statement
                sheet2.cell(row=i, column=2).value = 'worked'
                found = True
                break
        
        # Go to the next page if the keyword was not found in the current page
        if not found:
            next_page = soup.find("a", {"aria-label": "Page 2"})
            if next_page:
                response = requests.get(f'https://www.google.com{next_page["href"]}')
                soup = BeautifulSoup(response.text, 'html.parser')
            else:
                break

# Save the workbook
workbook.save('search_data.xlsx')
